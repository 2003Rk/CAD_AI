"""Report generation — JSON, HTML, Excel, and Google Sheets reports for evaluation results."""

from __future__ import annotations

import json
import logging
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from jinja2 import Template
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import gspread
from google.oauth2.service_account import Credentials as SACredentials

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# JSON report
# ---------------------------------------------------------------------------

def generate_json_report(
    results: list[dict],
    output_path: Path,
    metadata: dict | None = None,
) -> Path:
    """Write evaluation results as a structured JSON report."""
    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "total_evaluations": len(results),
        "metadata": metadata or {},
        "summary": _compute_summary(results),
        "results": results,
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")
    logger.info("JSON report written to %s", output_path)
    return output_path


# ---------------------------------------------------------------------------
# HTML report
# ---------------------------------------------------------------------------

_HTML_TEMPLATE = Template("""\
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>CAD AI Evaluation Report</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: #f5f5f5; color: #333; padding: 2rem; }
  .container { max-width: 1100px; margin: 0 auto; }
  h1 { margin-bottom: 0.5rem; }
  .meta { color: #666; margin-bottom: 2rem; font-size: 0.9rem; }
  .summary { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 1rem; margin-bottom: 2rem; }
  .card { background: #fff; border-radius: 8px; padding: 1.5rem; box-shadow: 0 1px 3px rgba(0,0,0,.1); }
  .card h3 { font-size: 0.85rem; color: #888; text-transform: uppercase; letter-spacing: 0.5px; }
  .card .value { font-size: 2rem; font-weight: 700; margin-top: 0.25rem; }
  .score-good { color: #22c55e; }
  .score-mid { color: #f59e0b; }
  .score-low { color: #ef4444; }
  table { width: 100%; border-collapse: collapse; background: #fff; border-radius: 8px; overflow: hidden; box-shadow: 0 1px 3px rgba(0,0,0,.1); margin-bottom: 2rem; }
  th, td { padding: 0.75rem 1rem; text-align: left; border-bottom: 1px solid #eee; }
  th { background: #fafafa; font-size: 0.8rem; text-transform: uppercase; color: #888; }
  .bar { display: inline-block; height: 8px; border-radius: 4px; }
  .pattern-section { margin-bottom: 3rem; }
  .pattern-title { font-size: 1.2rem; margin-bottom: 1rem; padding-bottom: 0.5rem; border-bottom: 2px solid #e5e5e5; }
</style>
</head>
<body>
<div class="container">
  <h1>CAD AI Evaluation Report</h1>
  <p class="meta">Generated: {{ generated_at }} &middot; Total evaluations: {{ total }}</p>

  <div class="summary">
    <div class="card">
      <h3>Overall Average</h3>
      <div class="value {{ 'score-good' if avg_overall >= 70 else ('score-mid' if avg_overall >= 40 else 'score-low') }}">{{ "%.1f"|format(avg_overall) }}%</div>
    </div>
    <div class="card">
      <h3>Geometry</h3>
      <div class="value">{{ "%.1f"|format(avg_geometry) }}%</div>
    </div>
    <div class="card">
      <h3>Structure</h3>
      <div class="value">{{ "%.1f"|format(avg_structure) }}%</div>
    </div>
    <div class="card">
      <h3>Dimensions</h3>
      <div class="value">{{ "%.1f"|format(avg_dimension) }}%</div>
    </div>
    <div class="card">
      <h3>Metadata</h3>
      <div class="value">{{ "%.1f"|format(avg_metadata) }}%</div>
    </div>
    <div class="card">
      <h3>Success Rate</h3>
      <div class="value {{ 'score-good' if success_rate >= 70 else ('score-mid' if success_rate >= 40 else 'score-low') }}">{{ "%.0f"|format(success_rate) }}%</div>
    </div>
  </div>

  {% for pattern_name, items in by_pattern.items() %}
  <div class="pattern-section">
    <h2 class="pattern-title">Prompt Pattern: {{ pattern_name }}</h2>
    <table>
      <thead>
        <tr>
          <th>Drawing</th>
          <th>Overall</th>
          <th>Geometry</th>
          <th>Structure</th>
          <th>Dimensions</th>
          <th>Metadata</th>
          <th>Status</th>
        </tr>
      </thead>
      <tbody>
      {% for r in items %}
        <tr>
          <td>{{ r.get("drawing_name", "N/A") }}</td>
          <td><strong>{{ "%.1f"|format(r.get("overall_score", 0)) }}</strong></td>
          <td>{{ "%.1f"|format(r.get("geometry_score", 0)) }}</td>
          <td>{{ "%.1f"|format(r.get("structure_score", 0)) }}</td>
          <td>{{ "%.1f"|format(r.get("dimension_score", 0)) }}</td>
          <td>{{ "%.1f"|format(r.get("metadata_score", 0)) }}</td>
          <td>{{ "✅" if r.get("success") else "❌ " + r.get("error", "")[:50] }}</td>
        </tr>
      {% endfor %}
      </tbody>
    </table>
  </div>
  {% endfor %}

</div>
</body>
</html>
""")


def generate_html_report(
    results: list[dict],
    output_path: Path,
    metadata: dict | None = None,
) -> Path:
    """Render evaluation results as an HTML report."""
    summary = _compute_summary(results)

    by_pattern: dict[str, list[dict]] = {}
    for r in results:
        pname = r.get("pattern_name", "Unknown")
        by_pattern.setdefault(pname, []).append(r)

    successful = [r for r in results if r.get("success")]
    success_rate = (len(successful) / len(results) * 100) if results else 0

    html = _HTML_TEMPLATE.render(
        generated_at=datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        total=len(results),
        avg_overall=summary.get("avg_overall", 0),
        avg_geometry=summary.get("avg_geometry", 0),
        avg_structure=summary.get("avg_structure", 0),
        avg_dimension=summary.get("avg_dimension", 0),
        avg_metadata=summary.get("avg_metadata", 0),
        success_rate=success_rate,
        by_pattern=by_pattern,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(html, encoding="utf-8")
    logger.info("HTML report written to %s", output_path)
    return output_path


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

def generate_excel_report(
    results: list[dict],
    output_path: Path,
) -> Path:
    """Write evaluation results to an Excel spreadsheet.

    Layout mirrors the reference spreadsheet:
    - One row per unique drawing/image
    - Columns: Dataset | ファイル名 | 元データ | 入力画像 | [thumbnail] |
               then for each prompt pattern: Prompt | 出力ファイル | 再現度評価
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    drawings: dict[str, dict[str, dict]] = {}
    pattern_names: list[str] = []
    for r in results:
        name = r.get("drawing_name", "")
        pname = r.get("pattern_name", "")
        drawings.setdefault(name, {})[pname] = r
        if pname and pname not in pattern_names:
            pattern_names.append(pname)

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook has no active worksheet")
    ws.title = "Evaluation Results"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F4F8F")
    sub_fill    = PatternFill("solid", fgColor="4472C4")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    base_cols = 5  # Dataset | ファイル名 | 元データ | 入力画像 | thumbnail
    cols_per_pattern = 3

    # Row 1 group headers
    for col, label in enumerate(["Dataset", "ファイル名", "元データ", "入力画像", ""], start=1):
        cell = ws.cell(1, col, label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for i, pname in enumerate(pattern_names):
        col = base_cols + 1 + i * cols_per_pattern
        cell = ws.cell(1, col, pname)
        cell.font = header_font
        cell.fill = sub_fill
        cell.alignment = center_align
        ws.merge_cells(
            start_row=1, start_column=col,
            end_row=1, end_column=col + cols_per_pattern - 1,
        )

    # Row 2 sub-headers
    for c in range(1, base_cols + 1):
        ws.cell(2, c).fill = sub_fill
        ws.cell(2, c).font = Font(bold=True, color="FFFFFF")
        ws.cell(2, c).alignment = center_align

    for i in range(len(pattern_names)):
        base = base_cols + 1 + i * cols_per_pattern
        for j, label in enumerate(["Prompt", "出力ファイル", "再現度評価"]):
            cell = ws.cell(2, base + j, label)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = sub_fill
            cell.alignment = center_align

    ws.column_dimensions[get_column_letter(5)].width = 10
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    for row_idx, (drawing_name, pattern_map) in enumerate(drawings.items(), start=3):
        any_result = next(iter(pattern_map.values()))
        img_path_str = any_result.get("image", "") or any_result.get("image_path", "")
        ref_path_str = any_result.get("reference", "")
        img_path = Path(img_path_str) if img_path_str else None
        dataset = img_path.parent.name if img_path else ""
        # Derive reference DXF name from drawing_name when not present in result
        ref_name = Path(ref_path_str).name if ref_path_str else f"{drawing_name}.dxf"

        ws.cell(row_idx, 1, dataset).alignment = center_align
        ws.cell(row_idx, 2, drawing_name).alignment = center_align
        ws.cell(row_idx, 3, ref_name).alignment = center_align
        ws.cell(row_idx, 4, img_path.name if img_path else "").alignment = center_align

        if img_path and img_path.exists():
            try:
                xl_img = XLImage(str(img_path))
                xl_img.width = 72
                xl_img.height = 54
                ws.add_image(xl_img, f"E{row_idx}")
                ws.row_dimensions[row_idx].height = 45
            except Exception:
                ws.row_dimensions[row_idx].height = 15
        else:
            ws.row_dimensions[row_idx].height = 15

        for i, pname in enumerate(pattern_names):
            base = base_cols + 1 + i * cols_per_pattern
            r = pattern_map.get(pname, {})
            out_file = Path(r.get("generated", "") or r.get("output", "") or r.get("output_path", "")).name
            score = r.get("overall_score")
            score_val = round(score / 100, 2) if score is not None else ""

            ws.cell(row_idx, base,     pname).alignment = center_align
            ws.cell(row_idx, base + 1, out_file).alignment = center_align
            score_cell = ws.cell(row_idx, base + 2, score_val)
            score_cell.alignment = center_align
            if isinstance(score_val, float):
                score_cell.number_format = "0.0"

        total_cols = base_cols + len(pattern_names) * cols_per_pattern
        for c in range(1, total_cols + 1):
            ws.cell(row_idx, c).border = border

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 24
    for i in range(len(pattern_names)):
        base = base_cols + 1 + i * cols_per_pattern
        ws.column_dimensions[get_column_letter(base)].width = 20
        ws.column_dimensions[get_column_letter(base + 1)].width = 26
        ws.column_dimensions[get_column_letter(base + 2)].width = 12

    ws.freeze_panes = "A3"

    wb.save(str(output_path))
    logger.info("Excel report written to %s", output_path)
    return output_path


# ---------------------------------------------------------------------------
# Google Sheets report
# ---------------------------------------------------------------------------

_SHEETS_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.file",
]


def generate_sheets_report(
    results: list[dict],
    credentials_path: Path,
    sheet_id: str,
    tab_name: str = "Evaluation Results",
    drive_folder_id: str | None = None,
) -> str:
    """Write evaluation results to a Google Sheet.

    Flat layout — one row per result (60 rows for 20 drawings × 3 patterns):
    - Row 1: header
    - Row 2+: Dataset | ファイル名 | 元データ | 入力画像 | Prompt | 出力ファイル | Status | 再現度評価 | 配点 | Geometry | Structure | Dimension | Metadata

    Returns the URL of the updated spreadsheet.
    """
    credentials_path = Path(credentials_path)
    creds = SACredentials.from_service_account_file(
        str(credentials_path), scopes=_SHEETS_SCOPES
    )
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(sheet_id)

    try:
        ws = sh.worksheet(tab_name)
        try:
            sh.batch_update({"requests": [{"unmergeCells": {"range": {
                "sheetId": ws.id,
                "startRowIndex": 0, "endRowIndex": 1000,
                "startColumnIndex": 0, "endColumnIndex": 50,
            }}}]})
        except Exception:
            pass
        ws.clear()
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=tab_name, rows=200, cols=20)

    header = [
        "Dataset",
        "ファイル名",
        "元データ",
        "入力画像",
        "Prompt",
        "出力ファイル",
        "Status",
        "再現度評価",
        "配点",
        "Geometry (40%)",
        "Structure (30%)",
        "Dimension (20%)",
        "Metadata (10%)",
    ]
    total_cols = len(header)

    # Load pre-uploaded image IDs (from upload_images_to_drive.py OAuth script)
    _ids_file = credentials_path.parent / "drive_image_ids.json"
    preloaded_image_ids: dict[str, str] = {}
    if _ids_file.exists():
        try:
            preloaded_image_ids = json.loads(_ids_file.read_text())
            logger.info("Loaded %d pre-uploaded image IDs from %s", len(preloaded_image_ids), _ids_file)
        except Exception:
            pass

    data_rows: list[list] = []

    # Sort results: by drawing_name then pattern_id for consistent ordering
    sorted_results = sorted(results, key=lambda r: (r.get("drawing_name", ""), r.get("pattern_id", 0)))

    for r in sorted_results:
        drawing_name = r.get("drawing_name", "")
        img_path_str = r.get("image", "") or r.get("image_path", "")
        ref_path_str = r.get("reference", "")
        img_path = Path(img_path_str) if img_path_str else None
        dataset = img_path.parent.name if img_path else ""
        ref_name = Path(ref_path_str).name if ref_path_str else f"{drawing_name}.dxf"

        image_formula = ""
        img_name = img_path.name if img_path else ""

        # Priority 1: use pre-uploaded Drive IDs from upload_images_to_drive.py
        if img_name and img_name in preloaded_image_ids:
            file_id = preloaded_image_ids[img_name]
            # Use custom sizing so line drawings remain visible inside the cell.
            url = f"https://lh3.googleusercontent.com/d/{file_id}=s1024"
            image_formula = f'=IMAGE("{url}", 4, 110, 180)'
        else:
            image_formula = img_name  # fallback: just show filename

        out_file = Path(r.get("generated", "") or r.get("output", "") or r.get("output_path", "")).name
        status = "OK" if r.get("success") else "NG"
        geometry_score = round(float(r.get("geometry_score", 0) or 0), 2)
        structure_score = round(float(r.get("structure_score", 0) or 0), 2)
        dimension_score = round(float(r.get("dimension_score", 0) or 0), 2)
        metadata_score = round(float(r.get("metadata_score", 0) or 0), 2)
        score = r.get("overall_score")
        # Keep score column numeric for every row, even when generation/comparison failed.
        score_val = round(score / 100, 2) if score is not None else 0
        weight_formula = "Geometry 40% + Structure 30% + Dimension 20% + Metadata 10%"

        data_rows.append([
            dataset,
            drawing_name,
            ref_name,
            image_formula,
            r.get("pattern_name", ""),
            out_file,
            status,
            score_val,
            weight_formula,
            geometry_score,
            structure_score,
            dimension_score,
            metadata_score,
        ])

    ws.update([header] + data_rows, value_input_option=gspread.utils.ValueInputOption.user_entered)

    header_color = {"red": 0.184, "green": 0.310, "blue": 0.561}
    white = {"red": 1.0, "green": 1.0, "blue": 1.0}

    # Alternate row colors: light blue for even data rows
    alt_color = {"red": 0.906, "green": 0.933, "blue": 0.976}

    # Fetch existing banded ranges for this worksheet and delete them first
    delete_banding_requests: list[dict] = []
    try:
        spreadsheet_meta = gc.http_client.request(
            "get",
            f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}",
            params={"fields": "sheets(properties(sheetId),bandedRanges)"},
        ).json()
        for sheet_meta in spreadsheet_meta.get("sheets", []):
            if sheet_meta.get("properties", {}).get("sheetId") == ws.id:
                for br in sheet_meta.get("bandedRanges", []):
                    delete_banding_requests.append(
                        {"deleteBanding": {"bandedRangeId": br["bandedRangeId"]}}
                    )
    except Exception:
        pass  # best-effort

    banding_request = {
        "addBanding": {
            "bandedRange": {
                "range": {
                    "sheetId": ws.id,
                    "startRowIndex": 1,
                    "endRowIndex": 1 + len(data_rows),
                    "startColumnIndex": 0,
                    "endColumnIndex": total_cols,
                },
                "rowProperties": {
                    "headerColor": {"red": 0.267, "green": 0.447, "blue": 0.769},
                    "firstBandColor": {"red": 1.0, "green": 1.0, "blue": 1.0},
                    "secondBandColor": alt_color,
                },
            }
        }
    }

    # Delete existing banding first (separate call, must precede addBanding)
    if delete_banding_requests:
        sh.batch_update({"requests": delete_banding_requests})

    sh.batch_update({"requests": [
        # Header row formatting
        {
            "repeatCell": {
                "range": {
                    "sheetId": ws.id,
                    "startRowIndex": 0, "endRowIndex": 1,
                    "startColumnIndex": 0, "endColumnIndex": total_cols,
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": header_color,
                        "textFormat": {"bold": True, "foregroundColor": white},
                        "horizontalAlignment": "CENTER",
                        "verticalAlignment": "MIDDLE",
                    }
                },
                "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment)",
            }
        },
        # Center-align score column (G = index 6)
        {
            "repeatCell": {
                "range": {
                    "sheetId": ws.id,
                    "startRowIndex": 1,
                    "endRowIndex": 1 + len(data_rows),
                    "startColumnIndex": 6, "endColumnIndex": 7,
                },
                "cell": {"userEnteredFormat": {"horizontalAlignment": "CENTER"}},
                "fields": "userEnteredFormat.horizontalAlignment",
            }
        },
        # Column widths
        {
            "updateDimensionProperties": {
                "range": {"sheetId": ws.id, "dimension": "COLUMNS", "startIndex": 0, "endIndex": 1},
                "properties": {"pixelSize": 120}, "fields": "pixelSize",
            }
        },
        {
            "updateDimensionProperties": {
                "range": {"sheetId": ws.id, "dimension": "COLUMNS", "startIndex": 1, "endIndex": 2},
                "properties": {"pixelSize": 200}, "fields": "pixelSize",
            }
        },
        {
            "updateDimensionProperties": {
                "range": {"sheetId": ws.id, "dimension": "COLUMNS", "startIndex": 2, "endIndex": 3},
                "properties": {"pixelSize": 180}, "fields": "pixelSize",
            }
        },
        {
            "updateDimensionProperties": {
                "range": {"sheetId": ws.id, "dimension": "COLUMNS", "startIndex": 3, "endIndex": 4},
                "properties": {"pixelSize": 180}, "fields": "pixelSize",
            }
        },
        {
            "updateDimensionProperties": {
                "range": {"sheetId": ws.id, "dimension": "COLUMNS", "startIndex": 4, "endIndex": 5},
                "properties": {"pixelSize": 200}, "fields": "pixelSize",
            }
        },
        {
            "updateDimensionProperties": {
                "range": {"sheetId": ws.id, "dimension": "COLUMNS", "startIndex": 5, "endIndex": 6},
                "properties": {"pixelSize": 200}, "fields": "pixelSize",
            }
        },
        {
            "updateDimensionProperties": {
                "range": {"sheetId": ws.id, "dimension": "COLUMNS", "startIndex": 6, "endIndex": 7},
                "properties": {"pixelSize": 100}, "fields": "pixelSize",
            }
        },
        # Row height for data rows
        {
            "updateDimensionProperties": {
                "range": {
                    "sheetId": ws.id, "dimension": "ROWS",
                    "startIndex": 1, "endIndex": 1 + len(data_rows),
                },
                "properties": {"pixelSize": 110},
                "fields": "pixelSize",
            }
        },
        # Freeze header row
        {
            "updateSheetProperties": {
                "properties": {
                    "sheetId": ws.id,
                    "gridProperties": {"frozenRowCount": 1},
                },
                "fields": "gridProperties.frozenRowCount",
            }
        },
        banding_request,
    ]})
    # Note: delete_banding_requests are appended after addBanding intentionally
    # — they must actually run first, so re-order:

    # Write the prompt definitions to a separate "Prompt list" tab
    _write_prompt_list_tab(sh)

    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}"
    logger.info("Google Sheets report written to %s (tab: %s)", url, tab_name)
    return url


# ---------------------------------------------------------------------------
# Prompt list tab
# ---------------------------------------------------------------------------

def _write_prompt_list_tab(sh: gspread.Spreadsheet, tab_name: str = "Prompt list") -> None:
    """Write the three prompt patterns to a dedicated tab.

    Layout matches the reference screenshot:
    - Row 1: header (プロンプト名 | プロンプト内容)
    - Row 2+: one row per pattern (name | system_prompt + user_prompt_template)
    """
    from src.prompts.patterns import PATTERN_STRUCTURED, PATTERN_STEPWISE, PATTERN_REFERENCE

    patterns = [PATTERN_STRUCTURED, PATTERN_STEPWISE, PATTERN_REFERENCE]

    try:
        ws = sh.worksheet(tab_name)
        try:
            sh.batch_update({"requests": [{"unmergeCells": {"range": {
                "sheetId": ws.id,
                "startRowIndex": 0, "endRowIndex": 200,
                "startColumnIndex": 0, "endColumnIndex": 10,
            }}}]})
        except Exception:
            pass
        ws.clear()
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=tab_name, rows=100, cols=10)

    header = ["プロンプト名", "プロンプト内容"]
    rows = [header]
    for p in patterns:
        full_text = p.system_prompt.strip() + "\n\n---\nUser template:\n" + p.user_prompt_template.strip()
        rows.append([p.name, full_text])

    ws.update(rows, value_input_option=gspread.utils.ValueInputOption.raw)

    header_color = {"red": 0.184, "green": 0.310, "blue": 0.561}
    white = {"red": 1.0, "green": 1.0, "blue": 1.0}

    sh.batch_update({"requests": [
        # Header row formatting
        {
            "repeatCell": {
                "range": {
                    "sheetId": ws.id,
                    "startRowIndex": 0, "endRowIndex": 1,
                    "startColumnIndex": 0, "endColumnIndex": 2,
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": header_color,
                        "textFormat": {"bold": True, "foregroundColor": white},
                        "horizontalAlignment": "CENTER",
                        "verticalAlignment": "MIDDLE",
                    }
                },
                "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment)",
            }
        },
        # Wrap text for all data rows
        {
            "repeatCell": {
                "range": {
                    "sheetId": ws.id,
                    "startRowIndex": 1, "endRowIndex": 1 + len(patterns),
                    "startColumnIndex": 0, "endColumnIndex": 2,
                },
                "cell": {
                    "userEnteredFormat": {
                        "wrapStrategy": "WRAP",
                        "verticalAlignment": "TOP",
                    }
                },
                "fields": "userEnteredFormat(wrapStrategy,verticalAlignment)",
            }
        },
        # Column A width (prompt name)
        {
            "updateDimensionProperties": {
                "range": {"sheetId": ws.id, "dimension": "COLUMNS", "startIndex": 0, "endIndex": 1},
                "properties": {"pixelSize": 180},
                "fields": "pixelSize",
            }
        },
        # Column B width (prompt content)
        {
            "updateDimensionProperties": {
                "range": {"sheetId": ws.id, "dimension": "COLUMNS", "startIndex": 1, "endIndex": 2},
                "properties": {"pixelSize": 700},
                "fields": "pixelSize",
            }
        },
        # Freeze header row
        {
            "updateSheetProperties": {
                "properties": {
                    "sheetId": ws.id,
                    "gridProperties": {"frozenRowCount": 1},
                },
                "fields": "gridProperties.frozenRowCount",
            }
        },
    ]})
    logger.info("Prompt list tab written to '%s'", tab_name)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _compute_summary(results: list[dict]) -> dict:
    """Compute average scores across all results."""
    total_count = len(results)
    success_count = sum(1 for r in results if r.get("success"))
    failed_count = total_count - success_count
    success_rate = (success_count / total_count * 100) if total_count else 0.0

    error_counter = Counter(
        str(r.get("error", "")).strip()
        for r in results
        if not r.get("success") and str(r.get("error", "")).strip()
    )

    scored = [r for r in results if r.get("success") and r.get("overall_score") is not None]
    if not scored:
        by_pattern: dict[str, dict] = {}
        for r in results:
            pname = str(r.get("pattern_name", "Unknown"))
            by_pattern.setdefault(pname, {"total": 0, "success": 0})
            by_pattern[pname]["total"] += 1
            if r.get("success"):
                by_pattern[pname]["success"] += 1
        for pname, stats in by_pattern.items():
            stats["success_rate"] = round((stats["success"] / stats["total"] * 100), 2) if stats["total"] else 0.0

        return {
            "avg_overall": 0, "avg_geometry": 0, "avg_structure": 0,
            "avg_dimension": 0, "avg_metadata": 0,
            "best_pattern": "N/A", "evaluated_count": 0,
            "total_count": total_count,
            "success_count": success_count,
            "failed_count": failed_count,
            "success_rate": success_rate,
            "by_pattern": by_pattern,
            "top_errors": dict(error_counter.most_common(5)),
        }

    def avg(key: str) -> float:
        return sum(r.get(key, 0) for r in scored) / len(scored)

    pattern_scores: dict[str, list[float]] = {}
    for r in scored:
        pn = r.get("pattern_name", "?")
        pattern_scores.setdefault(pn, []).append(r.get("overall_score", 0))
    best_pattern = max(pattern_scores, key=lambda k: sum(pattern_scores[k]) / len(pattern_scores[k]))

    by_pattern: dict[str, dict] = {}
    for r in results:
        pname = str(r.get("pattern_name", "Unknown"))
        by_pattern.setdefault(pname, {"total": 0, "success": 0, "scores": []})
        by_pattern[pname]["total"] += 1
        if r.get("success"):
            by_pattern[pname]["success"] += 1
        if r.get("overall_score") is not None:
            by_pattern[pname]["scores"].append(float(r.get("overall_score", 0)))

    for pname, stats in by_pattern.items():
        stats["success_rate"] = round((stats["success"] / stats["total"] * 100), 2) if stats["total"] else 0.0
        stats["avg_overall"] = round((sum(stats["scores"]) / len(stats["scores"])), 2) if stats["scores"] else 0.0
        del stats["scores"]

    return {
        "avg_overall": avg("overall_score"),
        "avg_geometry": avg("geometry_score"),
        "avg_structure": avg("structure_score"),
        "avg_dimension": avg("dimension_score"),
        "avg_metadata": avg("metadata_score"),
        "best_pattern": best_pattern,
        "evaluated_count": len(scored),
        "total_count": total_count,
        "success_count": success_count,
        "failed_count": failed_count,
        "success_rate": success_rate,
        "by_pattern": by_pattern,
        "top_errors": dict(error_counter.most_common(5)),
    }


